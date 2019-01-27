# -*- coding: utf-8 -*-

import os;
import sys;
import requests;
from PIL import Image;
from io import BytesIO;
from openpyxl import Workbook;
from openpyxl import load_workbook;
from matplotlib import pyplot as plt;

def greeting():
    print("this software(demo version) is used for USST ACC Library inventory");
    print("after 'name' on the screen, input the title of the book");
    print("after 'which one', choose one of them, within number 1~10");
    print("input EXACTLY 'quit!' to quit");
    print("don't forget the exclamation mark");
    print("remember to check the inventory afterwards");
    print("================================================");
    
def getImage(count,item):
    image_file=str(count)+".jpg";
    image_website=item["image"];
    r=requests.get(image_website);
            
    if(r.status_code==requests.codes.ok):
        image=Image.open(BytesIO(r.content));
        image=image.convert("RGB");
        image.save(image_file);
        
def showImage(count):
    for i in range(1,count): #[1,count-1]
        file_name=str(i)+".jpg";
        plt.subplot(3,3,i);
        img=plt.imread(file_name);
        plt.imshow(img);
        plt.title(file_name);
        plt.axis("off");
    plt.figure(figsize=(960,1));
    plt.show();
    
def removeImage(count):
     for i in range(1,count):
         file_name=str(i)+".jpg";
         os.remove(file_name);

def queryBook(bookname, page):
    payload={"q":bookname,"count":"9","start":page*9};
    
    r=requests.get("https://api.douban.com/v2/book/search",params=payload);
    r.raise_for_status();
    blist=r.json();
    ret=[];
    
    for item in blist["books"]:
        ret.append({"title":item["title"],"author":item["author"],
                    "publisher":item["publisher"],"isbn":item["isbn13"],
                    "image":item["image"]});
    #ISBN use only isbn13
    return ret;

def creatXlsx(fname):
    wb=Workbook();
    ws1=wb.active;
    ws1.title="books";
    ws1.append(["title","author","publisher","isbn"]);
    wb.save(filename=fname);
    
def appendXlsx(fname,title,author,publisher,isbn):
    if (os.access(fname,os.F_OK)==False):
        creatXlsx(fname);
    wb=load_workbook(filename=fname);
    ws=wb.active;
    ws.append([title,author,publisher,isbn]);
    wb.save(filename = fname);
    
if __name__ == '__main__':
    filename="inventory.xlsx";
    greeting();
    
    while(True):
        book_name=input("name:");
        if(book_name=="quit!"):
            sys.exit(0); #exit the programme
        
        book_infor=queryBook(book_name,0);  #book_infor is a list
        
        count=1;
        for item in book_infor:  #traverse the book_infor(list)
            print("number "+str(count));
            
            getImage(count,item);
                
            for key,value in item.items():  #traverse the item(dictionary)
            
                if(type(value)==list):  #for author member, the type is list
                    value=",".join(value);  #convert list to string
                print(key+":"+value);
                
            count+=1;
            print("\n",end="");
        
        showImage(count);
        removeImage(count);
        
        choice_num=int(input("which one:"));
        if(choice_num>=1 and choice_num<=len(book_infor)):
            title=book_infor[choice_num-1]["title"];
            author=book_infor[choice_num-1]["author"];
            author=",".join(author);  #convert list to string
            publisher=book_infor[choice_num-1]["publisher"];
            isbn=book_infor[choice_num-1]["isbn"];
            
            appendXlsx(filename,title,author,publisher,isbn);
        else:
            print("wrong!");